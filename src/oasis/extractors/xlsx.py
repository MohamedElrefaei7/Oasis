import logging
from pathlib import Path

import openpyxl

from oasis.extractors.base import stat_metadata
from oasis.models import ExtractedDocument

logger = logging.getLogger(__name__)


class XlsxExtractor:
    extensions: frozenset[str] = frozenset({".xlsx"})

    def extract(self, path: Path) -> ExtractedDocument | None:
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        except Exception:
            logger.warning("Failed to open XLSX %s", path, exc_info=True)
            return None

        # try/finally, not close-on-both-paths: read_only workbooks hold an open
        # file handle, and the previous shape closed it on the success path and
        # again in the handler — so a failure *after* the close double-closed,
        # and a failure between open and close on some paths leaked. One exit.
        try:
            lines: list[str] = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(sheet_name)
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        lines.append(row_text)

            text = "\n".join(lines)
            props = wb.properties
            title: str | None = props.title or None
            author: str | None = props.creator or None
            sheet_count = len(wb.sheetnames)

            return ExtractedDocument(
                path=path,
                text=text,
                metadata=stat_metadata(
                    path, title=title, author=author, page_count=sheet_count
                ),
            )
        except Exception:
            logger.warning("Failed to extract content from XLSX %s", path, exc_info=True)
            return None
        finally:
            wb.close()
